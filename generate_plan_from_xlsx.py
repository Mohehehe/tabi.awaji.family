import sys
from pathlib import Path
try:
    from openpyxl import load_workbook
except Exception as e:
    print('ERROR: openpyxl is required. Please run: pip install openpyxl')
    raise


def read_sheet(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = []
    for r in ws.iter_rows(values_only=True):
        # convert None to empty string
        rows.append(['' if c is None else str(c) for c in r])
    return rows


def guess_date_from_name(p: Path):
    name = p.stem
    # try to find 12月31日 pattern
    return name


def make_html(rows, title, out_path):
    # simple handwritten-style CSS using Google Fonts
    html = f'''<!doctype html>
<html lang="ja">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>{title}</title>
  <link href="https://fonts.googleapis.com/css2?family=Kosugi+Maru&display=swap" rel="stylesheet">
  <style>
    body{{font-family:'Kosugi Maru', sans-serif;background:linear-gradient(#fff,#f7fbff);padding:32px}}
    .paper{{max-width:800px;margin:0 auto;background:#fff;border:6px dashed #ffdede;border-radius:16px;padding:28px;box-shadow:0 6px 18px rgba(0,0,0,0.08)}}
    h1{{margin:0 0 12px;color:#d65a8d;text-align:center;font-size:32px}}
    .meta{{text-align:center;color:#666;margin-bottom:18px}}
    table{width:100%;border-collapse:collapse}
    td,th{padding:10px;border-bottom:1px dashed #eee}
    th{background:linear-gradient(#fff6f6,#fff);text-align:left;color:#b84d76}
    .time{width:110px;color:#b84d76;font-weight:700}
    .note{font-size:14px;color:#444}
    .emoji{font-size:20px;margin-right:8px}
    .footer{margin-top:18px;text-align:right;color:#999;font-size:13px}
  </style>
</head>
<body>
  <div class="paper">
    <h1>旅行のしおり — {title}</h1>
    <div class="meta">12月31日のスケジュール（Excel を元に自動生成）</div>
    <table>
      <thead>
        <tr>
'''

    # header detection: first row all non-empty and not numeric-only
    header = None
    if rows:
        first = rows[0]
        nonempty = sum(1 for c in first if c.strip()!='')
        if nonempty >= 2:
            header = first

    if header:
        for h in header:
            html += f'        <th>{h}</th>\n'
        html += '      </tr>\n      </thead>\n      <tbody>\n'
        data_rows = rows[1:]
    else:
        # generic columns: Time / Activity / Memo
        html += '        <th class="time">時間</th>\n        <th>内容</th>\n        <th>備考</th>\n      </tr>\n      </thead>\n      <tbody>\n'
        data_rows = rows

    # fill rows
    for r in data_rows:
        # ensure length
        while len(r) < 3:
            r.append('')
        t,a,n = r[0], r[1], ' / '.join([c for c in r[2:] if c.strip()!=''])
        emoji = '🕘' if t.strip() else '📍'
        html += f'        <tr>\n          <td class="time">{emoji} {t}</td>\n          <td class="note">{a}</td>\n          <td class="note">{n}</td>\n        </tr>\n'

    html += '''      </tbody>
    </table>
    <div class="footer">楽しい旅になりますように！✈️</div>
  </div>
</body>
</html>'''

    out_path.write_text(html, encoding='utf-8')


def main():
    if len(sys.argv) < 2:
        print('Usage: python generate_plan_from_xlsx.py path/to/12月31日.xlsx [out.html]')
        return
    xlsx = Path(sys.argv[1])
    if not xlsx.exists():
        print('ERROR: file not found', xlsx)
        return
    rows = read_sheet(xlsx)
    title = guess_date_from_name(xlsx)
    out = Path(sys.argv[2]) if len(sys.argv) >= 3 else xlsx.parent / 'plan_dec31.html'
    make_html(rows, title, out)
    print('Generated', out)


if __name__ == '__main__':
    main()
